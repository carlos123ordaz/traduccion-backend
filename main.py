from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse
import pandas as pd
from typing import Literal
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, range_boundaries
import os
import requests
from msal import ConfidentialClientApplication
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv

load_dotenv()

app = FastAPI()


TENANT_ID = os.getenv("TENANT_ID")
CLIENT_ID = os.getenv("CLIENT_ID")
CLIENT_SECRET = os.getenv("CLIENT_SECRET")

archivos_config = [
    {
        'unique_id': 'C36C72E3-D62D-4165-9042-5F0F16635B56',
        'nombre': '002_Compras_OCI.xlsx'
    },
    {
        'unique_id': 'BE104FCC-839F-49C1-A73B-CD0285A6858C',
        'nombre': 'Traduccion-Equipos.xlsx'
    },
]

download_dir = "./descargas"
os.makedirs(download_dir, exist_ok=True)

archivos_inicializados = False


def obtener_token():
    msal_app = ConfidentialClientApplication(
        CLIENT_ID,
        authority=f"https://login.microsoftonline.com/{TENANT_ID}",
        client_credential=CLIENT_SECRET
    )
    result = msal_app.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    )
    return result["access_token"]


def obtener_drive_id(token):
    headers = {'Authorization': f'Bearer {token}'}
    site_url = "https://graph.microsoft.com/v1.0/sites/corsusaadmin.sharepoint.com:/sites/logistica"
    site_response = requests.get(site_url, headers=headers)

    if site_response.status_code != 200:
        raise Exception(f"Error al obtener sitio: {site_response.status_code}")

    site_id = site_response.json()['id']
    drives_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives"
    drives_response = requests.get(drives_url, headers=headers)
    drives = drives_response.json()['value']

    document_drive = None
    for d in drives:
        if 'Documentos' in d['name']:
            document_drive = d
            break

    if not document_drive:
        document_drive = drives[0]

    return document_drive['id']


def descargar_archivo(config, drive_id, headers, download_dir):
    try:
        unique_id = config['unique_id']
        nombre = config['nombre']
        file_url = f"https://graph.microsoft.com/v1.0/drives/{drive_id}/items/{unique_id}"
        file_response = requests.get(file_url, headers=headers)

        if file_response.status_code != 200:
            return None, f"Error al obtener info del archivo {nombre}: {file_response.status_code}"

        file_data = file_response.json()
        download_url = file_data.get('@microsoft.graph.downloadUrl')

        if not download_url:
            return None, f"No se pudo obtener URL de descarga para {nombre}"

        file_content = requests.get(download_url)

        if file_content.status_code != 200:
            return None, f"Error al descargar {nombre}: {file_content.status_code}"

        ruta_archivo = os.path.join(download_dir, nombre)
        with open(ruta_archivo, 'wb') as f:
            f.write(file_content.content)

        return ruta_archivo, None

    except Exception as e:
        return None, f"Error descargando {config['nombre']}: {str(e)}"


def sincronizar_archivos():
    try:
        token = obtener_token()
        headers = {'Authorization': f'Bearer {token}'}
        drive_id = obtener_drive_id(token)

        archivos_descargados = []
        errores = []

        with ThreadPoolExecutor(max_workers=4) as executor:
            futures = {
                executor.submit(descargar_archivo, config, drive_id, headers, download_dir): config
                for config in archivos_config
            }

            for future in as_completed(futures):
                config = futures[future]
                try:
                    ruta, error = future.result()
                    if error:
                        errores.append(error)
                    else:
                        archivos_descargados.append(ruta)
                except Exception as e:
                    errores.append(
                        f"Excepción al procesar {config['nombre']}: {str(e)}")

        return {
            "exitoso": len(errores) == 0,
            "archivos_descargados": len(archivos_descargados),
            "total_archivos": len(archivos_config),
            "errores": errores
        }

    except Exception as e:
        return {
            "exitoso": False,
            "archivos_descargados": 0,
            "total_archivos": len(archivos_config),
            "errores": [str(e)]
        }


@app.on_event("startup")
async def startup_event():
    global archivos_inicializados
    print("Iniciando descarga inicial de archivos...")
    resultado = sincronizar_archivos()
    if resultado["exitoso"]:
        archivos_inicializados = True
        print(
            f"✅ Archivos inicializados: {resultado['archivos_descargados']}/{resultado['total_archivos']}")
    else:
        print(f"❌ Error en inicialización: {resultado['errores']}")


@app.get("/")
def read_root():
    return {
        "mensaje": "API de Traducción activa",
        "archivos_inicializados": archivos_inicializados
    }


@app.post("/sync")
def sincronizar():
    global archivos_inicializados
    resultado = sincronizar_archivos()
    if resultado["exitoso"]:
        archivos_inicializados = True
    return resultado


@app.get("/data")
def obtener_datos(type: Literal["embarque", "waybill"], value: str):
    if not archivos_inicializados:
        raise HTTPException(
            status_code=503,
            detail="Los archivos aún no están inicializados. Por favor espera o sincroniza manualmente."
        )

    try:
        ruta_traduccion = os.path.join(download_dir, "Traduccion-Equipos.xlsx")
        ruta_compras = os.path.join(download_dir, "002_Compras_OCI.xlsx")

        a = pd.read_excel(ruta_traduccion, sheet_name='Datos', skiprows=4)
        a = a.iloc[:, 1:]
        a['ID'] = a['Modelo'].astype(str) + '-' + a['Codigo'].astype(str)
        b = a[~a['ID'].duplicated()]
        b = b[~b['Modelo'].duplicated()]
        b['Codigo_Comercial'] = b['Modelo']
        c = pd.read_excel(ruta_compras, skiprows=2)
        df = pd.merge(c, b, how='left', on='Codigo_Comercial')
        df['PCU1'] = pd.to_numeric(df['PCU1'], errors='coerce').fillna(0)
        df['Cantidad'] = pd.to_numeric(
            df['Cantidad'], errors='coerce').fillna(0)
        df['Flete_US$'] = pd.to_numeric(
            df['Flete_US$'], errors='coerce').fillna(0)
        df['Sub Total'] = df['PCU1'] * df['Cantidad']
        df['Precio Total'] = df['Sub Total'] + df['Flete_US$']
        df = df[[
            'Item', 'Cantidad', 'Num_OC', 'Num_invoice', 'Codigo', 'Modelo', 'Descripcion',
            'Material', 'Uso', 'PaisOrigen', 'Moneda', 'PCU1', 'Sub Total',
            'Flete_US$', 'Precio Total', 'OperadorLogistico', 'Fecha_Invoice',
            'GrupoImportacion', 'Num_DocTransporte', 'RazonSocial_Proveedor',
            'Incoterm', 'Forma_Pago', 'Status_OCI', 'Marca'
        ]]
        new_columns = {
            'Cantidad': 'Cant',
            'Num_OC': 'Nº O.COMPRA',
            'Num_invoice': 'FACTURA',
            'PaisOrigen': 'País De Origen',
            'PCU1': 'Precio Unitario',
            'Flete_US$': 'Flete',
            'OperadorLogistico': 'TRANSPORTISTA',
            'Fecha_Invoice': 'FECHA FACTURA',
            'GrupoImportacion': 'Nº EMBARQUE',
            'Num_DocTransporte': 'Air Waybill',
            'RazonSocial_Proveedor': 'PROVEEDOR',
            'Incoterm': 'INCOTERM',
            'Forma_Pago': 'FORMA DE PAGO',
            'Status_OCI': 'ESTADO'
        }
        df.rename(columns=new_columns, inplace=True)
        if type == 'embarque':
            df = df[df['Nº EMBARQUE'].astype(
                str).str.contains(value, case=False, na=False)]
        else:
            df = df[df['Air Waybill'].astype(
                str).str.contains(value, case=False, na=False)]
        if df.empty:
            return {
                "data": [],
                "info": None,
                "mensaje": f"No se encontraron resultados para {type}: {value}"
            }
        df = df.replace({np.nan: None, np.inf: None, -np.inf: None})
        if 'FECHA FACTURA' in df.columns:
            df['FECHA FACTURA'] = df['FECHA FACTURA'].astype(str)
        primera_fila = df.iloc[0]
        info = {
            "TRANSPORTISTA": primera_fila['TRANSPORTISTA'] if pd.notna(primera_fila['TRANSPORTISTA']) else "",
            "FECHA FACTURA": str(primera_fila['FECHA FACTURA']) if pd.notna(primera_fila['FECHA FACTURA']) else "",
            "Nº EMBARQUE": primera_fila['Nº EMBARQUE'] if pd.notna(primera_fila['Nº EMBARQUE']) else "",
            "Air Waybill": primera_fila['Air Waybill'] if pd.notna(primera_fila['Air Waybill']) else "",
            "MARCA": primera_fila['Marca'] if pd.notna(primera_fila['Marca']) else "",
            "PROVEEDOR": primera_fila['PROVEEDOR'] if pd.notna(primera_fila['PROVEEDOR']) else "",
            "INCOTERM": primera_fila['INCOTERM'] if pd.notna(primera_fila['INCOTERM']) else "",
            "FORMA DE PAGO": primera_fila['FORMA DE PAGO'] if pd.notna(primera_fila['FORMA DE PAGO']) else "",
            "ESTADO": primera_fila['ESTADO'] if pd.notna(primera_fila['ESTADO']) else "",
        }
        return {
            "data": df.to_dict(orient='records'),
            "info": info
        }

    except FileNotFoundError as e:
        print(str(e))
        raise HTTPException(
            status_code=404, detail=f"Archivo no encontrado: {str(e)}")
    except KeyError as e:
        print(str(e))
        raise HTTPException(
            status_code=400, detail=f"Columna no encontrada en el archivo: {str(e)}")
    except Exception as e:
        print(str(e))
        raise HTTPException(
            status_code=500, detail=f"Error al procesar datos: {str(e)}")


@app.post("/export")
def exportar_datos(type: Literal["embarque", "waybill"], value: str):
    if not archivos_inicializados:
        raise HTTPException(
            status_code=503,
            detail="Los archivos aún no están inicializados. Por favor espera o sincroniza manualmente."
        )

    try:
        ruta_traduccion = os.path.join(download_dir, "Traduccion-Equipos.xlsx")
        ruta_compras = os.path.join(download_dir, "002_Compras_OCI.xlsx")
        ruta_plantilla = os.path.join(download_dir, "Plantilla.xlsx")

        a = pd.read_excel(ruta_traduccion, sheet_name='Datos', skiprows=4)
        a = a.iloc[:, 1:]
        a['ID'] = a['Modelo'].astype(str) + '-' + a['Codigo'].astype(str)
        b = a[~a['ID'].duplicated()]
        b = b[~b['Modelo'].duplicated()]
        b['Codigo_Comercial'] = b['Modelo']
        c = pd.read_excel(ruta_compras, skiprows=2)
        df = pd.merge(c, b, how='left', on='Codigo_Comercial')
        df['Sub Total'] = 0
        df['Precio Total'] = 0
        new_columns = {
            'Cantidad': 'Cant',
            'Num_OC': 'Nº O.COMPRA',
            'Num_invoice': 'FACTURA',
            'PaisOrigen': 'País De Origen',
            'PCU1': 'Precio Unitario',
            'Flete_US$': 'Flete',
            'OperadorLogistico': 'TRANSPORTISTA',
            'Fecha_Invoice': 'FECHA FACTURA',
            'GrupoImportacion': 'Nº EMBARQUE',
            'Num_DocTransporte': 'Air Waybill',
            'RazonSocial_Proveedor': 'PROVEEDOR',
            'Incoterm': 'INCOTERM',
            'Forma_Pago': 'FORMA DE PAGO',
            'Status_OCI': 'ESTADO'
        }
        df.rename(columns=new_columns, inplace=True)
        if type == 'embarque':
            df_filtered = df[df['Nº EMBARQUE'].astype(
                str).str.contains(value, case=False, na=False)]
        else:
            df_filtered = df[df['Air Waybill'].astype(
                str).str.contains(value, case=False, na=False)]

        primera_fila = df_filtered.iloc[0]
        info = {
            "TRANSPORTISTA": primera_fila['TRANSPORTISTA'] if pd.notna(primera_fila['TRANSPORTISTA']) else "",
            "FECHA FACTURA": str(primera_fila['FECHA FACTURA']) if pd.notna(primera_fila['FECHA FACTURA']) else "",
            "Nº EMBARQUE": primera_fila['Nº EMBARQUE'] if pd.notna(primera_fila['Nº EMBARQUE']) else "",
            "Air Waybill": primera_fila['Air Waybill'] if pd.notna(primera_fila['Air Waybill']) else "",
            "MARCA": primera_fila['Marca'] if pd.notna(primera_fila['Marca']) else "",
            "PROVEEDOR": primera_fila['PROVEEDOR'] if pd.notna(primera_fila['PROVEEDOR']) else "",
            "INCOTERM": primera_fila['INCOTERM'] if pd.notna(primera_fila['INCOTERM']) else "",
            "FORMA DE PAGO": primera_fila['FORMA DE PAGO'] if pd.notna(primera_fila['FORMA DE PAGO']) else "",
            "ESTADO": primera_fila['ESTADO'] if pd.notna(primera_fila['ESTADO']) else "",
        }

        df_filtered = df_filtered[[
            'Item', 'Cant', 'Nº O.COMPRA', 'FACTURA', 'Codigo', 'Modelo', 'Descripcion',
            'Material', 'Uso', 'País De Origen', 'Moneda', 'Precio Unitario', 'Sub Total',
            'Flete', 'Precio Total'
        ]]

        if df_filtered.empty:
            raise HTTPException(
                status_code=404, detail="No se encontraron datos para exportar")

        wb = load_workbook(ruta_plantilla)
        ws = wb.active

        table = ws.tables['Tabla24']
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        start_row = max_row + 1
        last_row = start_row - 1

        for row_idx, row in enumerate(
            dataframe_to_rows(df_filtered, index=False, header=False),
            start=start_row
        ):
            for col, value in enumerate(row, start=min_col):
                ws.cell(row=row_idx, column=col, value=value)
            last_row = row_idx

        table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{last_row}"

        for row in range(start_row, last_row + 1):
            ws[f'N{row}'] = f'=M{row}*C{row}'
            ws[f'P{row}'] = f'=O{row}+N{row}'

        ws['D2'] = info['TRANSPORTISTA']
        fecha = primera_fila['FECHA FACTURA']
        fecha_obj = pd.to_datetime(fecha)
        info['FECHA FACTURA'] = fecha_obj.strftime('%d/%m/%Y')
        ws['D4'] = info['Nº EMBARQUE']
        ws['D5'] = info['Air Waybill']

        ws['K2'] = info['PROVEEDOR']
        ws['K3'] = info['INCOTERM']
        ws['K4'] = info['FORMA DE PAGO']
        ws['K5'] = info['ESTADO']
        ws['K6'] = info['MARCA']

        output_file = 'Validación_xd.xlsx'
        wb.save(output_file)
        return FileResponse(
            path=output_file,
            filename=output_file,
            media_type='application/vnd.ms-excel.sheet.macroEnabled.12'
        )

    except FileNotFoundError as e:
        print(str(e))
        raise HTTPException(
            status_code=404, detail=f"Archivo no encontrado: {str(e)}")
    except Exception as e:
        print(str(e))
        raise HTTPException(
            status_code=500, detail=f"Error al exportar: {str(e)}")
